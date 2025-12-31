from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django import forms
from django.db.models import Count
from openpyxl import load_workbook
from .models import User, VideoUploadRecord, QueryLog


class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(label='Excel File', help_text='Upload Excel file with user data (.xlsx)')


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    list_display = ['username', 'phone', 'department', 'upload_count', 'query_count', 'is_active', 'is_staff', 'date_joined']
    list_filter = ['is_active', 'is_staff', 'department']
    search_fields = ['username', 'phone', 'department']
    ordering = ['-date_joined']

    fieldsets = BaseUserAdmin.fieldsets + (
        ('Extra Info', {'fields': ('phone', 'department')}),
    )
    add_fieldsets = BaseUserAdmin.add_fieldsets + (
        ('Extra Info', {'fields': ('phone', 'department')}),
    )

    change_list_template = 'admin/users/user/change_list.html'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(
            _upload_count=Count('uploads', distinct=True),
            _query_count=Count('query_logs', distinct=True)
        )

    def upload_count(self, obj):
        return obj._upload_count
    upload_count.short_description = 'Uploads'
    upload_count.admin_order_field = '_upload_count'

    def query_count(self, obj):
        return obj._query_count
    query_count.short_description = 'Queries'
    query_count.admin_order_field = '_query_count'

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='users_user_import_excel'),
            path('dashboard/', self.admin_site.admin_view(self.dashboard_view), name='users_dashboard'),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        from django.shortcuts import render, redirect
        from django.contrib import messages

        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    wb = load_workbook(excel_file)
                    ws = wb.active

                    created_count = 0
                    error_rows = []

                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not row[0]:
                            continue

                        username = str(row[0]).strip() if row[0] else ''
                        password = str(row[1]).strip() if len(row) > 1 and row[1] else 'password123'
                        phone = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                        department = str(row[3]).strip() if len(row) > 3 and row[3] else ''

                        if not username:
                            continue

                        if User.objects.filter(username=username).exists():
                            error_rows.append(f'Row {row_num}: Username {username} already exists')
                            continue

                        try:
                            User.objects.create_user(
                                username=username,
                                password=password,
                                phone=phone,
                                department=department,
                            )
                            created_count += 1
                        except Exception as e:
                            error_rows.append(f'Row {row_num}: {str(e)}')

                    if created_count > 0:
                        messages.success(request, f'Successfully imported {created_count} users')
                    if error_rows:
                        messages.warning(request, 'Some imports failed: ' + '; '.join(error_rows[:5]))

                    return redirect('..')
                except Exception as e:
                    messages.error(request, f'Import failed: {str(e)}')
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'title': 'Batch Import Users',
            'opts': self.model._meta,
        }
        return render(request, 'admin/users/user/import_excel.html', context)

    def dashboard_view(self, request):
        from django.shortcuts import render
        from django.db.models.functions import TruncDate
        from datetime import datetime, timedelta

        query_by_mode = QueryLog.objects.values('mode').annotate(count=Count('id')).order_by('-count')
        upload_by_brand = VideoUploadRecord.objects.values('brand').annotate(count=Count('id')).order_by('-count')

        seven_days_ago = datetime.now() - timedelta(days=7)
        daily_queries = QueryLog.objects.filter(created_at__gte=seven_days_ago).annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(count=Count('id')).order_by('date')

        top_uploaders = User.objects.annotate(
            upload_count=Count('uploads')
        ).filter(upload_count__gt=0).order_by('-upload_count')[:10]

        top_queriers = User.objects.annotate(
            query_count=Count('query_logs')
        ).filter(query_count__gt=0).order_by('-query_count')[:10]

        total_uploads = VideoUploadRecord.objects.count()
        total_queries = QueryLog.objects.count()

        context = {
            'title': 'Statistics Dashboard',
            'query_by_mode': list(query_by_mode),
            'upload_by_brand': list(upload_by_brand),
            'daily_queries': list(daily_queries),
            'top_uploaders': top_uploaders,
            'top_queriers': top_queriers,
            'total_uploads': total_uploads,
            'total_queries': total_queries,
            'opts': self.model._meta,
        }
        return render(request, 'admin/users/user/dashboard.html', context)


@admin.register(VideoUploadRecord)
class VideoUploadRecordAdmin(admin.ModelAdmin):
    list_display = ['user', 'brand', 'model', 'title', 'filename', 'created_at']
    list_filter = ['brand', 'created_at']
    search_fields = ['user__username', 'title', 'model', 'filename']
    ordering = ['-created_at']
    readonly_fields = ['user', 'brand', 'model', 'title', 'filename', 'file_path', 'created_at']


@admin.register(QueryLog)
class QueryLogAdmin(admin.ModelAdmin):
    list_display = ['user', 'mode', 'brand', 'keyword', 'created_at']
    list_filter = ['mode', 'brand', 'created_at']
    search_fields = ['user__username', 'keyword']
    ordering = ['-created_at']
    readonly_fields = ['user', 'mode', 'brand', 'keyword', 'created_at']
